"""Microsoft Excel document parser implementation."""

import logging
from typing import List, Dict, Any, Optional
import tempfile
import os
from pathlib import Path

from core.base import Parser, Document, ProcessingResult

logger = logging.getLogger(__name__)

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False
    pd = None  # Type hint placeholder
    logger.debug("pandas not available. Excel parsing will not be available.")

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    openpyxl = None  # Type hint placeholder
    logger.debug("openpyxl not available. Excel parsing may be limited.")


class ExcelParser(Parser):
    """Parser for Microsoft Excel documents (XLSX, XLS)."""
    
    def __init__(self, name: str = "ExcelParser", config: Optional[Dict[str, Any]] = None):
        super().__init__(name, config)
        config = config or {}
        self.parse_all_sheets = config.get("parse_all_sheets", True)
        self.max_rows_per_sheet = config.get("max_rows_per_sheet", 10000)
        self.include_formulas = config.get("include_formulas", False)
        self.treat_first_row_as_header = config.get("treat_first_row_as_header", True)
        self.separate_sheets_as_documents = config.get("separate_sheets_as_documents", True)
        self.extract_charts = config.get("extract_charts", False)
        self.skip_empty_rows = config.get("skip_empty_rows", True)
    
    def validate_config(self) -> bool:
        """Validate parser configuration."""
        if not HAS_PANDAS:
            logger.error("pandas package required for Excel parsing")
            return False
        return True
    
    def parse(self, source: str, **kwargs) -> ProcessingResult:
        """Parse Excel content into documents."""
        errors = []
        documents = []
        
        if not HAS_PANDAS:
            logger.error("pandas package required for Excel parsing")
            errors.append({"error": "pandas package not installed", "source": source})
            return ProcessingResult(documents=[], errors=errors)
        
        try:
            # Load Excel file directly from path
            documents = self._parse_excel_file(source, source=source, **kwargs)
            
        except Exception as e:
            logger.error(f"Error parsing Excel: {e}")
            errors.append({"error": str(e), "source": source})
            
        return ProcessingResult(documents=documents, errors=errors)
    
    def _parse_excel_file(self, file_path: str, **kwargs) -> List[Document]:
        """Parse the Excel file."""
        source = kwargs.get('source', 'unknown')
        documents = []
        
        try:
            # Read Excel file
            if self.parse_all_sheets:
                # Get all sheet names
                excel_file = pd.ExcelFile(file_path)
                sheet_names = excel_file.sheet_names
            else:
                # Just read the first sheet
                sheet_names = [0]
            
            for sheet_name in sheet_names:
                try:
                    # Read sheet
                    df = pd.read_excel(
                        file_path, 
                        sheet_name=sheet_name,
                        nrows=self.max_rows_per_sheet,
                        header=0 if self.treat_first_row_as_header else None
                    )
                    
                    if df.empty:
                        logger.debug(f"Skipping empty sheet: {sheet_name}")
                        continue
                    
                    # Create document for this sheet
                    if self.separate_sheets_as_documents:
                        doc = self._create_sheet_document(df, sheet_name, source)
                        documents.append(doc)
                    else:
                        # Add to combined document (handled later)
                        pass
                
                except Exception as e:
                    logger.warning(f"Error parsing sheet {sheet_name}: {e}")
                    continue
            
            # If not separating sheets, create combined document
            if not self.separate_sheets_as_documents and documents:
                combined_doc = self._create_combined_document(documents, source)
                documents = [combined_doc]
        
        except Exception as e:
            logger.error(f"Error reading Excel file: {e}")
            return []
        
        return documents
    
    def _create_sheet_document(self, df, sheet_name: str, source: str) -> Document:
        """Create a document from a single Excel sheet."""
        # Convert DataFrame to text content
        content = self._dataframe_to_text(df)
        
        # Create metadata
        metadata = {
            "type": "excel_sheet",
            "source": source,
            "sheet_name": str(sheet_name),
            "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }
        
        # Add DataFrame statistics
        metadata.update({
            "rows": len(df),
            "columns": len(df.columns),
            "column_names": df.columns.tolist(),
            "data_types": df.dtypes.astype(str).to_dict(),
            "has_headers": self.treat_first_row_as_header,
            "empty_cells": df.isnull().sum().sum(),
            "numeric_columns": df.select_dtypes(include=['number']).columns.tolist(),
            "text_columns": df.select_dtypes(include=['object']).columns.tolist()
        })
        
        # Add sample data
        if len(df) > 0:
            sample_rows = min(5, len(df))
            sample_data = df.head(sample_rows).to_dict('records')
            metadata["sample_data"] = sample_data
        
        # Extract additional insights
        if self._has_numeric_data(df):
            metadata["numeric_summary"] = self._get_numeric_summary(df)
        
        doc_id = f"excel_sheet_{sheet_name}_{hash(content) % 10000}"
        
        return Document(
            id=doc_id,
            content=content,
            metadata=metadata,
            source=source
        )
    
    def _create_combined_document(self, sheet_documents: List[Document], source: str) -> Document:
        """Create a combined document from multiple sheets."""
        # Combine content from all sheets
        combined_content = []
        combined_metadata = {
            "type": "excel_workbook",
            "source": source,
            "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "sheets": []
        }
        
        total_rows = 0
        total_columns = 0
        
        for doc in sheet_documents:
            combined_content.append(f"Sheet: {doc.metadata.get('sheet_name', 'Unknown')}")
            combined_content.append(doc.content)
            combined_content.append("")  # Separator
            
            # Aggregate metadata
            sheet_info = {
                "name": doc.metadata.get('sheet_name', 'Unknown'),
                "rows": doc.metadata.get('rows', 0),
                "columns": doc.metadata.get('columns', 0),
                "column_names": doc.metadata.get('column_names', [])
            }
            combined_metadata["sheets"].append(sheet_info)
            
            total_rows += doc.metadata.get('rows', 0)
            total_columns = max(total_columns, doc.metadata.get('columns', 0))
        
        combined_metadata.update({
            "total_sheets": len(sheet_documents),
            "total_rows": total_rows,
            "max_columns": total_columns
        })
        
        content = '\n'.join(combined_content)
        doc_id = f"excel_workbook_{hash(content) % 10000}"
        
        return Document(
            id=doc_id,
            content=content,
            metadata=combined_metadata,
            source=source
        )
    
    def _dataframe_to_text(self, df) -> str:
        """Convert DataFrame to readable text format."""
        if df.empty:
            return "Empty spreadsheet"
        
        lines = []
        
        # Add column headers if they exist
        if self.treat_first_row_as_header and hasattr(df, 'columns'):
            headers = [str(col) for col in df.columns]
            lines.append("Columns: " + " | ".join(headers))
            lines.append("")
        
        # Add data rows
        for idx, row in df.iterrows():
            if self.skip_empty_rows and row.isnull().all():
                continue
            
            # Convert row to string representation
            row_data = []
            for col_name, value in row.items():
                if pd.isnull(value):
                    row_data.append("N/A")
                else:
                    row_data.append(str(value))
            
            lines.append(" | ".join(row_data))
        
        return '\n'.join(lines)
    
    def _has_numeric_data(self, df) -> bool:
        """Check if DataFrame has numeric columns."""
        return len(df.select_dtypes(include=['number']).columns) > 0
    
    def _get_numeric_summary(self, df) -> Dict[str, Any]:
        """Get summary statistics for numeric columns."""
        numeric_df = df.select_dtypes(include=['number'])
        if numeric_df.empty:
            return {}
        
        summary = {}
        for col in numeric_df.columns:
            col_data = numeric_df[col].dropna()
            if len(col_data) > 0:
                summary[str(col)] = {
                    "count": len(col_data),
                    "mean": float(col_data.mean()),
                    "min": float(col_data.min()),
                    "max": float(col_data.max()),
                    "std": float(col_data.std()) if len(col_data) > 1 else 0.0
                }
        
        return summary
    
    def _extract_charts_info(self, file_path: str) -> List[Dict[str, Any]]:
        """Extract information about charts (requires openpyxl)."""
        if not HAS_OPENPYXL:
            return []
        
        charts_info = []
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path)
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                if hasattr(sheet, '_charts'):
                    for i, chart in enumerate(sheet._charts):
                        charts_info.append({
                            "sheet": sheet_name,
                            "chart_index": i,
                            "chart_type": type(chart).__name__,
                            "title": getattr(chart, 'title', None)
                        })
        except Exception as e:
            logger.debug(f"Could not extract chart info: {e}")
        
        return charts_info
    
    @classmethod
    def get_supported_extensions(cls) -> List[str]:
        """Get supported file extensions."""
        return ['.xlsx', '.xls', '.xlsm']
    
    @classmethod
    def get_description(cls) -> str:
        """Get parser description."""
        return "Microsoft Excel document parser that extracts data from spreadsheets, including multiple sheets and metadata."